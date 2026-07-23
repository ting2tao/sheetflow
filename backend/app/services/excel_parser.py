"""Excel file parser using openpyxl.

Reads .xlsx files and extracts:
- Cell values
- Font styles (name, size, bold, italic, color)
- Background/fill colors
- Borders
- Alignment
- Merged cells
- Column widths
- Row heights
"""

from dataclasses import dataclass, field
from typing import List, Optional, Dict, Any
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string


@dataclass
class CellStyle:
    """Style information for a single cell."""
    font_name: Optional[str] = None
    font_size: Optional[float] = None
    font_bold: bool = False
    font_italic: bool = False
    font_color: Optional[str] = None
    bg_color: Optional[str] = None
    border_top: Optional[str] = None
    border_bottom: Optional[str] = None
    border_left: Optional[str] = None
    border_right: Optional[str] = None
    align_horizontal: Optional[str] = None
    align_vertical: Optional[str] = None
    wrap_text: bool = False


@dataclass
class Cell:
    """A single cell with value and style."""
    value: Any = None
    style: CellStyle = field(default_factory=CellStyle)
    row: int = 0
    col: int = 0


@dataclass
class MergedCell:
    """Represents a merged cell range."""
    min_row: int
    max_row: int
    min_col: int
    max_col: int


@dataclass
class SheetModel:
    """Parsed sheet data."""
    name: str
    rows: List[List[Cell]] = field(default_factory=list)
    merged_cells: List[MergedCell] = field(default_factory=list)
    column_widths: Dict[int, float] = field(default_factory=dict)
    row_heights: Dict[int, float] = field(default_factory=dict)
    max_row: int = 0
    max_col: int = 0


@dataclass
class WorkbookModel:
    """Parsed workbook data."""
    sheets: List[SheetModel] = field(default_factory=list)
    active_sheet_index: int = 0


def _parse_color(color) -> Optional[str]:
    """Convert openpyxl color to CSS hex string."""
    if color is None:
        return None
    if color.type == 'rgb' and color.rgb:
        rgb = str(color.rgb)
        if len(rgb) == 8:  # ARGB format
            return f"#{rgb[2:]}"
        elif len(rgb) == 6:
            return f"#{rgb}"
    if color.type == 'theme':
        # Theme colors - use reasonable defaults
        theme_map = {
            0: "#FFFFFF", 1: "#000000", 2: "#E7E6E6", 3: "#44546A",
            4: "#4472C4", 5: "#ED7D31", 6: "#A5A5A5", 7: "#FFC000",
            8: "#5B9BD5", 9: "#70AD47",
        }
        return theme_map.get(color.theme, None)
    return None


def _parse_border(border) -> Optional[str]:
    """Convert openpyxl border to CSS border string."""
    if border is None or border.style is None:
        return None
    style_map = {
        'thin': '1px solid',
        'medium': '2px solid',
        'thick': '3px solid',
        'dashed': '1px dashed',
        'dotted': '1px dotted',
        'double': '3px double',
    }
    css_style = style_map.get(border.style, '1px solid')
    color = _parse_color(border.color) if border.color else '#000000'
    return f"{css_style} {color or '#000000'}"


def _parse_cell_style(cell) -> CellStyle:
    """Extract style from an openpyxl cell."""
    style = CellStyle()

    # Font
    if cell.font:
        style.font_name = cell.font.name
        style.font_size = cell.font.size
        style.font_bold = cell.font.bold or False
        style.font_italic = cell.font.italic or False
        style.font_color = _parse_color(cell.font.color)

    # Fill/Background
    if cell.fill and cell.fill.patternType:
        if cell.fill.patternType == 'solid':
            style.bg_color = _parse_color(cell.fill.fgColor)

    # Borders
    if cell.border:
        style.border_top = _parse_border(cell.border.top)
        style.border_bottom = _parse_border(cell.border.bottom)
        style.border_left = _parse_border(cell.border.left)
        style.border_right = _parse_border(cell.border.right)

    # Alignment
    if cell.alignment:
        style.align_horizontal = cell.alignment.horizontal
        style.align_vertical = cell.alignment.vertical
        style.wrap_text = cell.alignment.wrap_text or False

    return style


def _get_column_widths(worksheet) -> Dict[int, float]:
    """Get column widths from worksheet."""
    widths = {}
    for col_letter, dim in worksheet.column_dimensions.items():
        if dim.width:
            try:
                col_idx = column_index_from_string(col_letter)
                widths[col_idx] = dim.width
            except ValueError:
                continue
    return widths


def _get_row_heights(worksheet) -> Dict[int, float]:
    """Get row heights from worksheet."""
    heights = {}
    for row_num, dim in worksheet.row_dimensions.items():
        if dim.height:
            heights[row_num] = dim.height
    return heights


def get_sheet_list(file_path: str) -> list:
    """Quickly get sheet names and dimensions without parsing all data.

    Args:
        file_path: Path to the .xlsx file

    Returns:
        List of dicts with index, name, rows, columns
    """
    wb = load_workbook(file_path, read_only=True, data_only=True)
    sheets = []

    for i, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        # In read_only mode, max_row and max_column might be None
        max_row = ws.max_row if ws.max_row is not None else 0
        max_col = ws.max_column if ws.max_column is not None else 0
        sheets.append({
            "index": i,
            "name": sheet_name,
            "rows": max_row,
            "columns": max_col,
        })

    wb.close()
    return sheets


def parse_excel(file_path: str, sheet_indices: list = None) -> WorkbookModel:
    """Parse an Excel file and return structured data.

    Args:
        file_path: Path to the .xlsx file
        sheet_indices: List of sheet indices to parse. None means parse all.

    Returns:
        WorkbookModel with all sheets, cells, styles, and merge info
    """
    wb = load_workbook(file_path, data_only=True)
    model = WorkbookModel()

    # Determine which sheets to parse
    if sheet_indices is not None:
        sheet_names = [
            wb.sheetnames[i]
            for i in sheet_indices
            if 0 <= i < len(wb.sheetnames)
        ]
    else:
        sheet_names = wb.sheetnames

    for sheet_name in sheet_names:
        ws = wb[sheet_name]

        sheet = SheetModel(
            name=sheet_name,
            column_widths=_get_column_widths(ws),
            row_heights=_get_row_heights(ws),
            max_row=ws.max_row or 0,
            max_col=ws.max_column or 0,
        )

        # Parse merged cells
        for merge_range in ws.merged_cells.ranges:
            sheet.merged_cells.append(MergedCell(
                min_row=merge_range.min_row,
                max_row=merge_range.max_row,
                min_col=merge_range.min_col,
                max_col=merge_range.max_col,
            ))

        # Parse all cells
        for row_idx in range(1, (ws.max_row or 0) + 1):
            row_cells = []
            for col_idx in range(1, (ws.max_column or 0) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                parsed_cell = Cell(
                    value=cell.value,
                    style=_parse_cell_style(cell),
                    row=row_idx,
                    col=col_idx,
                )
                row_cells.append(parsed_cell)
            sheet.rows.append(row_cells)

        model.sheets.append(sheet)

    wb.close()
    return model


def get_sheet_data(model: WorkbookModel, sheet_index: int = 0) -> SheetModel:
    """Get a specific sheet from the workbook model."""
    if 0 <= sheet_index < len(model.sheets):
        return model.sheets[sheet_index]
    return model.sheets[0] if model.sheets else SheetModel(name="Empty")
